"""無 overfit 實驗：兩階段抽取（OCR 表格 → 讀文字抽尺寸）。
階段一 GPT-4o 看圖忠實 OCR Package Outline 表成 markdown；階段二讀文字套通用規則抽 L/W/H。
測 V23 錯的尺寸顆，比對正確 specbook。"""
import os, sys, json, base64
sys.stdout.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)
os.environ["BACKEND"] = "azure_gpt4o"
from dotenv import load_dotenv
load_dotenv()
import fitz
import config
from llm_backend import get_backend
from reevaluate import cmp_numeric

# V23 尺寸有錯的顆（CD4148WTP/BAS16/DFLS160/MSB30M/BAS21H/BAV99W/MBR15U150）
TEST = ["CD4148WTP", "BAS16", "BAS21H", "DFLS160", "MSB30M", "BAV99W", "MBR15U150"]

import openpyxl
def load_spec():
    wb = openpyxl.load_workbook(config.SPECBOOK_PATH)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        pn = rec.get("Part Number")
        if pn:
            out[pn] = rec
    return out

spec = load_spec()
backend = get_backend("azure_gpt4o")


def raster(pdf_path, zoom=5.0, maxp=6):
    doc = fitz.open(str(pdf_path))
    out = []
    mat = fitz.Matrix(zoom, zoom)
    for i in range(min(doc.page_count, maxp)):
        out.append(base64.b64encode(doc[i].get_pixmap(matrix=mat).tobytes("png")).decode("ascii"))
    doc.close()
    return out


STAGE1 = """看下方 datasheet 截圖，找 Package Outline / Mechanical Dimensions / Body Dimensions 表
（不是 Recommended Land Pattern / Footprint / Tape and Reel）。
把該表每個本體尺寸符號（A / D / E 及變體 D1/D2/E1/E2/H_E 等）那列的數值，忠實 OCR 抄成 JSON。
不要判讀、不要挑選，照抄。
輸出 JSON 格式：{{"rows": [{{"symbol": "D", "min": <數值或null>, "nom": <數值或null>, "max": <數值或null>}}, ...]}}"""

STAGE2 = """下方是某電子元件的 Package Outline 尺寸表（JSON，symbol 對 min/nom/max）。請依通用規則抽長寬高，輸出 JSON：
- D → Maximum Length，E → Maximum Width，A → Maximum Height（硬性符號對應，不論數值大小、不交換）
- 每個值取該列的 MAX 欄（最右欄），不要取 MIN / NOM / TYP
- 只能用 A / D / E 三個主符號，禁止用 D1/D2/E1/E2/H_E/L/L1/H 等變體
- 單位若為 inch 乘 25.4

尺寸表：
{TABLE}

只輸出 JSON：{{"Maximum Length (mm)": <D的MAX數值或null>, "Maximum Width (mm)": <E的MAX數值或null>, "Maximum Height (mm)": <A的MAX數值或null>}}"""


def parse(t):
    t = t.strip()
    if "```" in t:
        import re
        m = re.search(r"\{.*\}", t, re.DOTALL)
        if m:
            t = m.group(0)
    s, e = t.find("{"), t.rfind("}")
    return json.loads(t[s:e + 1]) if s >= 0 else {}


print(f"兩階段實驗 — 測 {len(TEST)} 顆\n", flush=True)
hit = 0
total = 0
for pn in TEST:
    pdf = config.DATASHEETS_DIR / config.PART_TO_PDF[pn]
    imgs = raster(pdf)
    table = backend.call_multimodal(STAGE1, imgs, timeout=config.TIMEOUT)
    try:
        ans = parse(backend.call(STAGE2.format(TABLE=table), timeout=config.TIMEOUT))
    except Exception as ex:
        print(f"[{pn}] stage2 parse fail: {ex}", flush=True)
        continue
    exp = spec[pn]
    parts = []
    for f, k in [("Maximum Length (mm)", "L"), ("Maximum Width (mm)", "W"), ("Maximum Height (mm)", "H")]:
        ok = cmp_numeric(ans.get(f), exp.get(f), config.NUMERIC_TOLERANCE)
        total += 1
        hit += int(ok)
        parts.append(f"{k}={'OK' if ok else 'X'}({ans.get(f)}/{exp.get(f)})")
    print(f"[{pn}] {'  '.join(parts)}", flush=True)

print(f"\n尺寸格命中: {hit}/{total} = {hit/total*100:.0f}%", flush=True)
