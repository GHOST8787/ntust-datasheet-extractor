"""無 overfit 實驗：用「含引腳最大外形」通用規則重抽尺寸（符合欄位名 Maximum）。
這是通用領域定義（不看特定 part 答案），不是 overfit。GPT-4o multimodal 跑全 10 顆，eval 尺寸欄。"""
import os, sys, json, base64
sys.stdout.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)
os.environ["BACKEND"] = "azure_gpt4o"
from dotenv import load_dotenv
load_dotenv()
import fitz
import openpyxl
import config
from llm_backend import get_backend
from reevaluate import cmp_numeric


def load_spec():
    wb = openpyxl.load_workbook(config.SPECBOOK_PATH)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        if rec.get("Part Number"):
            out[rec["Part Number"]] = rec
    return out


spec = load_spec()
backend = get_backend("azure_gpt4o")


def raster(pdf_path, zoom=5.0, maxp=6):
    doc = fitz.open(str(pdf_path))
    out = []
    mat = fitz.Matrix(zoom, zoom)
    for i in range(min(doc.page_count, maxp)):
        out.append(base64.b64encode(doc[i].get_pixmap(matrix=mat).tobytes("png")).decode("ascii"))
    doc.close()
    return out


PROMPT = """看 datasheet 截圖的 Package Outline / Mechanical Dimensions 表（不是 Land Pattern / Footprint / Tape），
抽這顆零件的「最大外形尺寸」。

# 關鍵定義（通用，適用所有封裝）
- **Maximum Length / Maximum Width = 含引腳的最大外形尺寸（overall, lead-to-lead span）**，不是 body 本體尺寸。
  - 例：SOT-23 的 E（lead span ~2.5）是含引腳外形，E1（body ~1.3）是本體 → Width 取 E 那個含引腳的較大值。
  - 表中若同時有 body 符號與含引腳符號（如 E vs E1、或 D vs H_E / lead span），取「含引腳/外形較大」那個。
- **Maximum Height = 封裝總高 A**。
- 三個值都取該列的 **MAX 欄（最右欄）**，不是 MIN / NOM / TYP。
- Length 取較長軸方向、Width 取較短軸方向的含引腳外形（依圖示方向）。
- 單位若為 inch，乘 25.4。

# 輸出 JSON
{{"Maximum Length (mm)": <數值或null>, "Maximum Width (mm)": <數值或null>, "Maximum Height (mm)": <數值或null>}}
只輸出 JSON。"""


def parse(t):
    t = t.strip()
    import re
    m = re.search(r"\{.*\}", t, re.DOTALL)
    return json.loads(m.group(0)) if m else {}


print("含引腳最大外形規則 — 全 10 顆尺寸測\n", flush=True)
hit = 0
total = 0
for pn, pdf_fn in config.PART_TO_PDF.items():
    pdf = config.DATASHEETS_DIR / pdf_fn
    imgs = raster(pdf)
    try:
        ans = parse(backend.call_multimodal(PROMPT, imgs, timeout=config.TIMEOUT))
    except Exception as ex:
        print(f"[{pn}] FAIL: {ex}", flush=True)
        total += 3
        continue
    exp = spec[pn]
    parts = []
    for f, k in [("Maximum Length (mm)", "L"), ("Maximum Width (mm)", "W"), ("Maximum Height (mm)", "H")]:
        ok = cmp_numeric(ans.get(f), exp.get(f), config.NUMERIC_TOLERANCE)
        total += 1
        hit += int(ok)
        parts.append(f"{k}={'OK' if ok else 'X'}({ans.get(f)}/{exp.get(f)})")
    print(f"[{pn}] {'  '.join(parts)}", flush=True)

print(f"\n尺寸格命中: {hit}/{total} = {hit/total*100:.0f}%  (V23 baseline: 17/30=57%)", flush=True)
