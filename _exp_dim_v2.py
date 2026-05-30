"""零 overfit 衝刺：尺寸三欄(L/W/H)抽取精度提升。

根因(已用 datasheet 佐證)：V23/overall 尺寸失分 = VLM 在 Package Outline 表抽錯軸
  —— 把 body 符號(如 E1)當外形、把別軸數值(如 D)當寬。不是天花板、不是 GT 矛盾。

方法(全通用，不看這 10 份答案)：
  1. 一次 multimodal 呼叫，強制兩步：
     步驟一 — 忠實列出表中「每個」符號的 min/nom/max + 它量哪個軸 + 含引腳還是本體
     步驟二 — 依「含引腳最大外形 + 取 Max 欄」物理規則對應 L/W/H
  2. self-consistency：每顆跑 N 次，逐欄取多數(對應 VLM 一致性限制)
規則只叫模型「正確地看」，不直接設值、不點名 part、不照答案反推。

測全 10 顆尺寸三欄，對 specbook。
"""
import os, sys, json, base64, re
from collections import Counter
sys.stdout.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)
os.environ["BACKEND"] = "azure_gpt4o"
from dotenv import load_dotenv
load_dotenv()
import fitz
import openpyxl
import config
from llm_backend import get_backend
from reevaluate import cmp_numeric

N_SAMPLES = 3


def load_spec():
    wb = openpyxl.load_workbook(config.SPECBOOK_PATH)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        if rec.get("Part Number"):
            out[rec["Part Number"]] = rec
    return out


spec = load_spec()
backend = get_backend("azure_gpt4o")


def raster(pdf_path, zoom=5.0, maxp=6):
    doc = fitz.open(str(pdf_path))
    out = []
    mat = fitz.Matrix(zoom, zoom)
    for i in range(min(doc.page_count, maxp)):
        out.append(base64.b64encode(doc[i].get_pixmap(matrix=mat).tobytes("png")).decode("ascii"))
    doc.close()
    return out


PROMPT = """你在讀電子元件 datasheet 的 Package Outline / Mechanical Dimensions 表
(不是 Recommended Land Pattern / Footprint / Suggested Solder Pad / Tape and Reel)。
務必照下面兩步做：

【步驟一】忠實列表，不要judge：把該機構尺寸表中「每一個」符號照抄成清單，
每個給 min / nom / max 三個值(該欄沒有就 null)，並標註：
- axis：它量的是哪個方向(長軸 / 短軸 / 高度)
- kind：它是「含引腳/外形」還是「僅本體」
提示：SOT/SOD 類常有 E(含引腳 lead span) 與 E1(本體寬)；DO/SMA/SMB 類常有 E(含引腳全長) 與 E1(本體長)。

【步驟二】依通用物理定義對應(取最大外形)：
- Maximum Length = 最長軸方向「含引腳最大外形」符號的 Max 值。
  若同一方向同時有本體與含引腳兩個符號，取含引腳那個(較大)，不要選 body-only 符號。
- Maximum Width  = 次長軸方向「含引腳最大外形」符號的 Max 值(同原則)。
- Maximum Height = 封裝總高的 Max 值。
- 三者都取該符號的 Max 欄(最右/最大)，不要取 min / nom / typ。
- 單位若為 inch，乘 25.4。

輸出 JSON(先 symbols 後 result)：
{{"symbols": [{{"sym": "D", "axis": "長軸", "kind": "含引腳", "min": null, "nom": null, "max": null}}],
"Maximum Length (mm)": <數值>, "Maximum Width (mm)": <數值>, "Maximum Height (mm)": <數值>}}
只輸出 JSON。"""


def parse(t):
    t = t.strip()
    m = re.search(r"\{.*\}", t, re.DOTALL)
    if not m:
        return {}
    try:
        return json.loads(m.group(0))
    except Exception:
        return {}


def majority(vals):
    vals = [v for v in vals if v is not None]
    if not vals:
        return None
    c = Counter(str(v) for v in vals)
    top = c.most_common(1)[0][0]
    for v in vals:
        if str(v) == top:
            return v
    return vals[0]


DIM_FIELDS = [("Maximum Length (mm)", "L"), ("Maximum Width (mm)", "W"), ("Maximum Height (mm)", "H")]

print(f"尺寸抽取 v2(結構化+物理規則+self-consistency x{N_SAMPLES})— 全 10 顆\n", flush=True)
hit = 0
total = 0
for pn, pdf_fn in config.PART_TO_PDF.items():
    pdf = config.DATASHEETS_DIR / pdf_fn
    imgs = raster(pdf)
    samples = {f: [] for f, _ in DIM_FIELDS}
    for _ in range(N_SAMPLES):
        try:
            ans = parse(backend.call_multimodal(PROMPT, imgs, timeout=config.TIMEOUT))
        except Exception:
            continue
        for f, _ in DIM_FIELDS:
            samples[f].append(ans.get(f))
    exp = spec[pn]
    parts = []
    for f, k in DIM_FIELDS:
        val = majority(samples[f])
        ok = cmp_numeric(val, exp.get(f), config.NUMERIC_TOLERANCE)
        total += 1
        hit += int(ok)
        parts.append(f"{k}={'OK' if ok else 'X'}({val}/{exp.get(f)})")
    print(f"[{pn}] {'  '.join(parts)}", flush=True)

print(f"\n尺寸格命中: {hit}/{total} = {hit/total*100:.0f}%  (V23: 17/30=57%, overall規則: 21/30=70%)", flush=True)
