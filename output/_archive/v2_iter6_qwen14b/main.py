"""V2 主程式：PDF → LLM → 比對 → Excel 報告"""
import json
import re
import sys
import time
from pathlib import Path
from typing import Dict, List, Tuple

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import requests

import config
from prompts import build_extraction_prompt
from validators import validate_and_fix

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)
    sys.stderr.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)
except Exception:
    pass


# ===================================================================
# 1. PDF 抽取（pdfplumber 文字 + 表格，三段式截斷防 LLM context 爆掉）
# ===================================================================

def extract_pdf_text(pdf_path: Path) -> str:
    pages: List[Dict] = []
    with pdfplumber.open(pdf_path) as pdf:
        for idx, page in enumerate(pdf.pages):
            if idx >= config.MAX_PDF_PAGES:
                break
            text = page.extract_text() or ""
            try:
                tables = page.extract_tables() or []
            except Exception:
                tables = []
            table_chunks = []
            for tbl in tables:
                rows = []
                for row in tbl:
                    cells = [str(c).strip() if c is not None else "" for c in row]
                    rows.append(" | ".join(cells))
                if rows:
                    table_chunks.append("\n".join(rows))
            tbl_text = "\n\n".join(table_chunks)
            combined = (text + " " + tbl_text).lower()
            has_kw = any(kw.lower() in combined for kw in config.PDF_KEYWORDS)
            pages.append({
                "page_num": idx + 1,
                "text": text,
                "tables": tbl_text,
                "has_keyword": has_kw,
            })

    def join(selected):
        chunks = []
        for p in selected:
            chunks.append(f"--- Page {p['page_num']} ---")
            chunks.append(p["text"])
            if p["tables"]:
                chunks.append("\n[Tables]")
                chunks.append(p["tables"])
            chunks.append("")
        return "\n".join(chunks)

    full = join(pages)
    if len(full) > config.MAX_PDF_CHARS:
        # 字元爆 → 只留第 1 頁 + 含關鍵字頁
        kw_only = [p for p in pages if p["page_num"] == 1 or p["has_keyword"]]
        full = join(kw_only)
        if len(full) > config.MAX_PDF_CHARS:
            full = full[:config.MAX_PDF_CHARS] + "\n... [TRUNCATED]"
    return full


# ===================================================================
# 2. LLM 呼叫（Ollama HTTP API）
# ===================================================================

def call_ollama(prompt: str) -> str:
    payload = {
        "model": config.OLLAMA_MODEL,
        "prompt": prompt,
        "stream": False,
        "format": "json",  # logits constraint 強制 JSON
        "options": {
            "temperature": config.TEMPERATURE,
            "num_ctx": config.NUM_CTX,
        },
    }
    resp = requests.post(config.OLLAMA_URL, json=payload, timeout=config.TIMEOUT)
    resp.raise_for_status()
    return resp.json().get("response", "")


def extract_json(text: str) -> Dict:
    text = text.strip()
    if text.startswith("```"):
        text = text[3:]
        if text.lower().startswith("json"):
            text = text[4:]
        if text.endswith("```"):
            text = text[:-3]
        text = text.strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    s, e = text.find("{"), text.rfind("}")
    if s == -1 or e == -1 or e <= s:
        raise ValueError(f"No JSON found: {text[:200]!r}")
    return json.loads(text[s:e + 1])


def call_with_retry(prompt: str) -> Dict:
    last = None
    for attempt in range(config.MAX_RETRIES):
        try:
            return extract_json(call_ollama(prompt))
        except (json.JSONDecodeError, ValueError, requests.RequestException) as e:
            last = e
            if attempt < config.MAX_RETRIES - 1:
                wait = 2 ** attempt
                print(f"      [retry {attempt + 1}] {e}; sleep {wait}s")
                time.sleep(wait)
    raise RuntimeError(f"LLM failed after {config.MAX_RETRIES} retries: {last}")


# ===================================================================
# 3. 比對（與 V1 相同：數值容差 / 字串完全 / token 集合）
# ===================================================================

def load_specbook(path: Path) -> Dict[str, Dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    spec = {}
    for r in rows[1:]:
        if not r or r[0] is None:
            continue
        rec = dict(zip(headers, r))
        pn = str(rec.get("Part Number", "")).strip()
        if pn:
            spec[pn] = rec
    return spec


def _norm_token(t: str) -> str:
    s = t.lower()
    s = s.replace("μ", "u").replace("µ", "u")
    s = re.sub(r"\b(ta|tj|vr|ir|if|vf|tc|tamb)\s*=\s*\+?", "", s)
    s = re.sub(r"\b(ma|a|v|w)(dc|ac|pk|rms)\b", r"\1", s)
    s = s.replace(",", "").replace("，", "").replace("、", "").replace("+", "").replace("°", "")
    s = re.sub(r"\s+", "", s)

    def strip0(m):
        n = m.group(0)
        if "." in n:
            n = n.rstrip("0").rstrip(".")
        return n

    s = re.sub(r"\d+\.\d+", strip0, s)
    return s


def split_text(v) -> set:
    if v is None or not isinstance(v, str):
        return set()
    return {_norm_token(p) for p in re.split(r"[、]", v) if p.strip()}


def cmp_numeric(x, e) -> bool:
    if x is None or e is None:
        return x == e
    try:
        ef, xf = float(e), float(x)
    except (ValueError, TypeError):
        return str(x).strip() == str(e).strip()
    if ef == 0:
        return abs(xf) < 1e-6
    return abs(xf - ef) / abs(ef) <= config.NUMERIC_TOLERANCE


def cmp_exact(x, e) -> bool:
    if x is None or e is None:
        return x == e

    def norm(v):
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()

    return norm(x) == norm(e)


def cmp_text(x, e) -> Tuple[bool, float]:
    es = split_text(e)
    xs = split_text(x)
    if not es:
        return (not xs), (1.0 if not xs else 0.0)
    if not xs:
        return False, 0.0
    score = len(es & xs) / len(es)
    return score >= config.TEXT_MATCH_THRESHOLD, score


def evaluate(extracted: Dict, spec: Dict) -> Dict:
    out = {
        "overall": {"correct": 0, "total": 0, "accuracy": 0.0},
        "per_part": {},
        "cells": [],
    }
    for pn, exp_rec in spec.items():
        e = extracted.get(pn, {})
        pr = {"correct": 0, "total": 0, "accuracy": 0.0, "fields": {}}
        if "_error" in e:
            for f in config.FIELDS:
                d = {"expected": exp_rec.get(f), "extracted": None,
                     "match": False, "reason": "extraction failed"}
                pr["fields"][f] = d
                out["cells"].append({
                    "part": pn, "field": f,
                    "expected": exp_rec.get(f), "extracted": None, "match": False,
                })
            pr["total"] = len(config.FIELDS)
        else:
            for f in config.FIELDS:
                exp = exp_rec.get(f)
                ext = e.get(f)
                if f in config.NUMERIC_FIELDS:
                    ok = cmp_numeric(ext, exp)
                    d = {"expected": exp, "extracted": ext, "match": ok}
                elif f in config.TEXT_FIELDS:
                    ok, score = cmp_text(ext, exp)
                    d = {"expected": exp, "extracted": ext, "match": ok,
                         "token_score": round(score, 3)}
                else:
                    ok = cmp_exact(ext, exp)
                    d = {"expected": exp, "extracted": ext, "match": ok}
                pr["fields"][f] = d
                pr["total"] += 1
                if ok:
                    pr["correct"] += 1
                out["cells"].append({
                    "part": pn, "field": f,
                    "expected": exp, "extracted": ext, "match": ok,
                })
        pr["accuracy"] = pr["correct"] / pr["total"] if pr["total"] else 0
        out["per_part"][pn] = pr
        out["overall"]["correct"] += pr["correct"]
        out["overall"]["total"] += pr["total"]
    if out["overall"]["total"]:
        out["overall"]["accuracy"] = out["overall"]["correct"] / out["overall"]["total"]
    return out


# ===================================================================
# 4. Excel 輸出（3 個 sheet：Results / Comparison / Summary）
# ===================================================================

def write_excel(extracted: Dict, report: Dict, path: Path) -> None:
    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    green = PatternFill("solid", fgColor="D5F0D5")
    red = PatternFill("solid", fgColor="FBD5D5")

    # ---- Sheet 1: Results（10 顆 × 11 欄抽取結果）----
    ws1 = wb.active
    ws1.title = "Results"
    ws1.append(config.FIELDS)
    for c in ws1[1]:
        c.font = bold
        c.fill = header_fill
    for pn in config.PART_TO_PDF.keys():
        rec = extracted.get(pn, {})
        if "_error" in rec:
            row = [pn] + [None] * (len(config.FIELDS) - 1)
        else:
            row = [rec.get(f) for f in config.FIELDS]
            if not row[0]:
                row[0] = pn
            # 把 None 轉空字串，list/dict 轉 str（Excel 不接受）
            row = [v if isinstance(v, (str, int, float)) or v is None else str(v) for v in row]
        ws1.append(row)
    for i, f in enumerate(config.FIELDS, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = max(14, min(45, len(f) + 2))
    ws1.freeze_panes = "B2"

    # ---- Sheet 2: Comparison（每格逐一對標答）----
    ws2 = wb.create_sheet("Comparison")
    ws2.append(["Part Number", "Field", "Expected", "Extracted", "Match"])
    for c in ws2[1]:
        c.font = bold
        c.fill = header_fill
    for cell in report["cells"]:
        row = [
            cell["part"],
            cell["field"],
            "" if cell["expected"] is None else str(cell["expected"]),
            "" if cell["extracted"] is None else str(cell["extracted"]),
            "✓" if cell["match"] else "✗",
        ]
        ws2.append(row)
        cur = ws2.max_row
        fill = green if cell["match"] else red
        for col in range(1, 6):
            ws2.cell(row=cur, column=col).fill = fill
    for i, w in enumerate([14, 38, 50, 50, 8], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # ---- Sheet 3: Summary（per-part + overall）----
    ws3 = wb.create_sheet("Summary")
    ws3.append(["Part Number", "Correct", "Total", "Accuracy"])
    for c in ws3[1]:
        c.font = bold
        c.fill = header_fill
    for pn, pr in report["per_part"].items():
        ws3.append([pn, pr["correct"], pr["total"], f"{pr['accuracy'] * 100:.1f}%"])
    o = report["overall"]
    ws3.append([])
    ws3.append(["Overall", o["correct"], o["total"], f"{o['accuracy'] * 100:.1f}%"])
    last = ws3.max_row
    for col in range(1, 5):
        ws3.cell(row=last, column=col).font = bold
        ws3.cell(row=last, column=col).fill = PatternFill("solid", fgColor="FFE699")
    for i, w in enumerate([14, 10, 8, 12], start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"

    wb.save(path)


# ===================================================================
# 5. 主流程
# ===================================================================

def run() -> Dict:
    config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print("[1/4] 載入標準答案")
    spec = load_specbook(config.SPECBOOK_PATH)
    print(f"      共 {len(spec)} 顆元件")

    print(f"[2/4] PDF 抽取 + LLM 推理(model: {config.OLLAMA_MODEL})")
    extracted: Dict[str, Dict] = {}
    for pn, pdf_filename in config.PART_TO_PDF.items():
        pdf_path = config.DATASHEETS_DIR / pdf_filename
        if not pdf_path.exists():
            print(f"      [WARN] {pn}: 找不到 {pdf_filename}")
            extracted[pn] = {"_error": "PDF not found"}
            continue
        print(f"      - {pn}", flush=True)
        try:
            pdf_text = extract_pdf_text(pdf_path)
            prompt = build_extraction_prompt(pn, pdf_text)
            extracted[pn] = call_with_retry(prompt)
            # 後處理 validator
            extracted[pn], fixes = validate_and_fix(extracted[pn], pn)
            for fix in fixes:
                print(f"        [validator] {fix}", flush=True)
        except Exception as e:
            print(f"        [ERROR] {e}")
            extracted[pn] = {"_error": str(e)}

    # JSON dump 保留供 debug
    (config.OUTPUT_DIR / "results.json").write_text(
        json.dumps(extracted, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print("[3/4] 評估比對")
    report = evaluate(extracted, spec)

    print("[4/4] 寫出 Excel")
    excel_path = config.OUTPUT_DIR / "results_v2.xlsx"
    write_excel(extracted, report, excel_path)
    print(f"      → {excel_path}")

    o = report["overall"]
    print()
    print(f"=== 整體準確率：{o['correct']}/{o['total']} = {o['accuracy'] * 100:.1f}% ===")
    print()
    print("各元件：")
    for pn, pr in report["per_part"].items():
        print(f"  {pn:<15} {pr['correct']:>3}/{pr['total']:>3} = {pr['accuracy'] * 100:>5.1f}%")
    return report


if __name__ == "__main__":
    try:
        run()
    except KeyboardInterrupt:
        print("\nUser interrupt")
        sys.exit(130)
