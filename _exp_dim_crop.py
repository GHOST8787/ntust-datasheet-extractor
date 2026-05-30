"""零 overfit 衝刺(最後合法手法)：裁切 Package Outline 表高清餵 VLM + overall 單步物理規則。

純輸入品質改善：聚焦尺寸表區塊、減少整頁雜訊，讓 VLM 少抽錯軸。
- prompt 沿用 _exp_overall 已驗證 70% 的單步「含引腳最大外形 + Max 欄」(不用退步的兩步法)
- 單次呼叫(控制變因：只改「裁切 vs 整頁」這一個變數)
- crop miss 的顆 fallback 整頁，並標明來源
測全 10 顆尺寸三欄，對 specbook。對照 V23 57% / overall(整頁) 70%。
"""
import os, sys, json, base64, re
sys.stdout.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)
os.environ["BACKEND"] = "azure_gpt4o"
from dotenv import load_dotenv
load_dotenv()
import fitz
import openpyxl
import config
from llm_backend import get_backend
from reevaluate import cmp_numeric
from crop_extractor import get_dimension_crops


def load_spec():
    wb = openpyxl.load_workbook(config.SPECBOOK_PATH)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        if rec.get("Part Number"):
            out[rec["Part Number"]] = rec
    return out


spec = load_spec()
backend = get_backend("azure_gpt4o")


def raster_full(pdf_path, zoom=5.0, maxp=6):
    doc = fitz.open(str(pdf_path))
    out = []
    mat = fitz.Matrix(zoom, zoom)
    for i in range(min(doc.page_count, maxp)):
        out.append(base64.b64encode(doc[i].get_pixmap(matrix=mat).tobytes("png")).decode("ascii"))
    doc.close()
    return out


# 與 _exp_overall 完全相同的單步 prompt(已驗證整頁 70%)
PROMPT = """看 datasheet 截圖的 Package Outline / Mechanical Dimensions 表(不是 Land Pattern / Footprint / Tape),
抽這顆零件的「最大外形尺寸」。

# 關鍵定義(通用,適用所有封裝)
- **Maximum Length / Maximum Width = 含引腳的最大外形尺寸(overall, lead-to-lead span)**,不是 body 本體尺寸。
  - 例:SOT-23 的 E(lead span ~2.5)是含引腳外形,E1(body ~1.3)是本體 → Width 取 E 那個含引腳的較大值。
  - 表中若同時有 body 符號與含引腳符號(如 E vs E1、或 D vs H_E / lead span),取「含引腳/外形較大」那個。
- **Maximum Height = 封裝總高 A**。
- 三個值都取該列的 **MAX 欄(最右欄)**,不是 MIN / NOM / TYP。
- Length 取較長軸方向、Width 取較短軸方向的含引腳外形(依圖示方向)。
- 單位若為 inch,乘 25.4。

# 輸出 JSON
{{"Maximum Length (mm)": <數值或null>, "Maximum Width (mm)": <數值或null>, "Maximum Height (mm)": <數值或null>}}
只輸出 JSON。"""


def parse(t):
    t = t.strip()
    m = re.search(r"\{.*\}", t, re.DOTALL)
    return json.loads(m.group(0)) if m else {}


DIM_FIELDS = [("Maximum Length (mm)", "L"), ("Maximum Width (mm)", "W"), ("Maximum Height (mm)", "H")]

print("尺寸抽取 — 裁切尺寸表 + overall 單步規則 — 全 10 顆\n", flush=True)
hit = 0
total = 0
for pn, pdf_fn in config.PART_TO_PDF.items():
    pdf = config.DATASHEETS_DIR / pdf_fn
    crops = get_dimension_crops(pdf, zoom=6.0, max_crops=3)
    if crops:
        imgs = [b64 for _, b64 in crops]
        src = f"crop x{len(imgs)}"
    else:
        imgs = raster_full(pdf)
        src = "full(crop miss)"
    try:
        ans = parse(backend.call_multimodal(PROMPT, imgs, timeout=config.TIMEOUT))
    except Exception as ex:
        print(f"[{pn}] ({src}) FAIL: {ex}", flush=True)
        total += 3
        continue
    exp = spec[pn]
    parts = []
    for f, k in DIM_FIELDS:
        ok = cmp_numeric(ans.get(f), exp.get(f), config.NUMERIC_TOLERANCE)
        total += 1
        hit += int(ok)
        parts.append(f"{k}={'OK' if ok else 'X'}({ans.get(f)}/{exp.get(f)})")
    print(f"[{pn}] ({src}) {'  '.join(parts)}", flush=True)

print(f"\n尺寸格命中: {hit}/{total} = {hit/total*100:.0f}%  (V23: 17/30=57%, overall整頁: 21/30=70%)", flush=True)
