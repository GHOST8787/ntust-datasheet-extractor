"""零 overfit 衝刺：裁切 + overall 單步規則 + 兩個通用補強。
在 _exp_dim_crop(77%)基礎上加兩件純通用改善：
  1. 裁切失敗的顆才補通用標題關鍵字「DIMENSIONS」(本來成功的 8 顆不動，保住戰果)
  2. prompt 加通用規則：尺寸若為 nominal±tolerance 格式，Max = nominal + tolerance
     (例 1.55±0.1 → 1.65；任何 datasheet 的 ±公差都適用，不點名 part)
  3. Height 符號通用化(A / T / H 都可能是總高)
測全 10 顆尺寸三欄。對照 V23 57% / overall 整頁 70% / 裁切 77%。
"""
import os, sys, json, base64, re
sys.stdout.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)
os.environ["BACKEND"] = "azure_gpt4o"
from dotenv import load_dotenv
load_dotenv()
import fitz
import openpyxl
import config
from llm_backend import get_backend
from reevaluate import cmp_numeric
import crop_extractor

ORIG_KW = list(crop_extractor.CROP_ANCHOR_KEYWORDS)


def load_spec():
    wb = openpyxl.load_workbook(config.SPECBOOK_PATH)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        if rec.get("Part Number"):
            out[rec["Part Number"]] = rec
    return out


spec = load_spec()
backend = get_backend("azure_gpt4o")


def raster_full(pdf_path, zoom=5.0, maxp=6):
    doc = fitz.open(str(pdf_path))
    out = []
    mat = fitz.Matrix(zoom, zoom)
    for i in range(min(doc.page_count, maxp)):
        out.append(base64.b64encode(doc[i].get_pixmap(matrix=mat).tobytes("png")).decode("ascii"))
    doc.close()
    return out


def crops_default(pdf):
    crop_extractor.CROP_ANCHOR_KEYWORDS = ORIG_KW
    return crop_extractor.get_dimension_crops(pdf, zoom=6.0, max_crops=3)


def crops_extended(pdf):
    # 只在 default 失敗時用：補通用標題關鍵字 DIMENSIONS
    crop_extractor.CROP_ANCHOR_KEYWORDS = ORIG_KW + ["DIMENSIONS", "Dimensions"]
    try:
        return crop_extractor.get_dimension_crops(pdf, zoom=6.0, max_crops=3)
    finally:
        crop_extractor.CROP_ANCHOR_KEYWORDS = ORIG_KW


PROMPT = """看 datasheet 截圖的 Package Outline / Mechanical Dimensions / DIMENSIONS 表
(不是 Land Pattern / Footprint / Suggested Solder Pad / Tape),抽這顆零件的「最大外形尺寸」。

# 關鍵定義(通用,適用所有封裝)
- **Maximum Length / Maximum Width = 含引腳的最大外形尺寸(overall, lead-to-lead span)**,不是 body 本體尺寸。
  - 例:SOT-23 的 E(lead span ~2.5)是含引腳外形,E1(body ~1.3)是本體 → Width 取 E 那個含引腳的較大值。
  - 表中若同時有 body 符號與含引腳符號(如 E vs E1、或 D vs H_E / lead span),取「含引腳/外形較大」那個。
- **Maximum Height = 封裝總高**(常見符號 A,也可能標成 T 或 H)。
- 三個值都取該列的 **MAX 欄(最右/最大)**,不是 MIN / NOM / TYP。
- **若尺寸標成「中心值 ± 公差」格式(例如 1.55±0.1),Max = 中心值 + 公差(=1.65),不要只取中心值。**
- Length 取較長軸方向、Width 取較短軸方向的含引腳外形(依圖示方向)。
- 單位若為 inch,乘 25.4。

# 輸出 JSON
{{"Maximum Length (mm)": <數值或null>, "Maximum Width (mm)": <數值或null>, "Maximum Height (mm)": <數值或null>}}
只輸出 JSON。"""


def parse(t):
    t = t.strip()
    m = re.search(r"\{.*\}", t, re.DOTALL)
    return json.loads(m.group(0)) if m else {}


DIM_FIELDS = [("Maximum Length (mm)", "L"), ("Maximum Width (mm)", "W"), ("Maximum Height (mm)", "H")]

print("尺寸抽取 — 裁切(失敗補DIMENSIONS) + 單步規則 + ±公差規則 — 全 10 顆\n", flush=True)
hit = 0
total = 0
for pn, pdf_fn in config.PART_TO_PDF.items():
    pdf = config.DATASHEETS_DIR / pdf_fn
    crops = crops_default(pdf)
    if crops:
        src = f"crop x{len(crops)}"
    else:
        crops = crops_extended(pdf)
        if crops:
            src = f"crop+DIM x{len(crops)}"
        else:
            crops = None
    if crops:
        imgs = [b64 for _, b64 in crops]
    else:
        imgs = raster_full(pdf)
        src = "full(crop miss)"
    try:
        ans = parse(backend.call_multimodal(PROMPT, imgs, timeout=config.TIMEOUT))
    except Exception as ex:
        print(f"[{pn}] ({src}) FAIL: {ex}", flush=True)
        total += 3
        continue
    exp = spec[pn]
    parts = []
    for f, k in DIM_FIELDS:
        ok = cmp_numeric(ans.get(f), exp.get(f), config.NUMERIC_TOLERANCE)
        total += 1
        hit += int(ok)
        parts.append(f"{k}={'OK' if ok else 'X'}({ans.get(f)}/{exp.get(f)})")
    print(f"[{pn}] ({src}) {'  '.join(parts)}", flush=True)

print(f"\n尺寸格命中: {hit}/{total} = {hit/total*100:.0f}%  (V23 57% / overall 70% / 裁切 77%)", flush=True)
