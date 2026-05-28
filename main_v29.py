"""V2 主程式：PDF → LLM → 比對 → Excel 報告 → 自動歸檔

用法:
  # 1. 從 .env.example 拷成 .env 並填值
  # 2. .env 裡設 BACKEND=ollama / azure_gpt4o / azure_llama / azure_mistral
  # 3. python main.py

每次跑完會自動把結果歸檔到 output/_archive/v2_<backend_tag>/
"""
import json
import re
import shutil
import sys
import time
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# 讀 .env(必須在 import llm_backend 之前)
load_dotenv()

import os

import config
from prompts_v29 import (
    build_extraction_prompt,
    build_extraction_prompt_with_feedback,
    build_extraction_prompt_multimodal,
    build_extraction_prompt_with_feedback_multimodal,
    build_extraction_prompt_multimodal_cot,
    build_extraction_prompt_with_feedback_multimodal_cot,
)
from validators import validate_and_fix
from llm_backend import get_backend, LLMBackend
from pdf_extractor import get_pdf_extractor, PdfExtractor, MultiModalExtractor
from qa_critic import call_critic, get_critic_mode
from iteration_logger import IterationLogger
from output_schema import OPENAI_RESPONSE_FORMAT_STRICT

# V3/V4 雙 Agent 迭代上限（imlacha V7 同設定）
MAX_CRITIC_ROUNDS = 3


def get_use_strict_schema() -> bool:
    """V5 開關：環境變數 USE_STRICT_SCHEMA=1 啟用 Pydantic JSON Schema strict mode"""
    return os.environ.get("USE_STRICT_SCHEMA", "0").lower() in ("1", "true", "yes")


def get_use_cot() -> bool:
    """V6 開關：環境變數 USE_CHAIN_OF_THOUGHT=1 啟用強制 _image_reasoning 欄位"""
    return os.environ.get("USE_CHAIN_OF_THOUGHT", "0").lower() in ("1", "true", "yes")

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)
    sys.stderr.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)
except Exception:
    pass


# ===================================================================
# 1. PDF 抽取(透過 pdf_extractor 抽象層)
# ===================================================================
# (extract_pdf_text 已移到 pdf_extractor.py 抽象層,從 env 讀 PDF_EXTRACTOR 切換)


# ===================================================================
# 2. LLM 呼叫(透過 backend 抽象層)
# ===================================================================

def extract_json(text: str) -> Dict:
    text = text.strip()
    if text.startswith("```"):
        text = text[3:]
        if text.lower().startswith("json"):
            text = text[4:]
        if text.endswith("```"):
            text = text[:-3]
        text = text.strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    s, e = text.find("{"), text.rfind("}")
    if s == -1 or e == -1 or e <= s:
        raise ValueError(f"No JSON found: {text[:200]!r}")
    return json.loads(text[s:e + 1])


def call_with_retry(backend: LLMBackend, prompt: str) -> Dict:
    last = None
    for attempt in range(config.MAX_RETRIES):
        try:
            return extract_json(backend.call(prompt, timeout=config.TIMEOUT))
        except (json.JSONDecodeError, ValueError, Exception) as e:
            last = e
            if attempt < config.MAX_RETRIES - 1:
                wait = 2 ** attempt
                print(f"      [retry {attempt + 1}] {type(e).__name__}: {e}; sleep {wait}s")
                time.sleep(wait)
    raise RuntimeError(f"LLM failed after {config.MAX_RETRIES} retries: {last}")


# ===================================================================
# V3 雙 Agent 自我校正迴圈（imlacha V7 風格）
# ===================================================================

def run_dual_agent_loop(
    backend: LLMBackend,
    pn: str,
    pdf_text: str,
    logger: "IterationLogger" = None,
    images_b64: List[str] = None,
    multimodal: bool = False,
) -> Tuple[Dict, List[Dict]]:
    """雙 Agent 迭代抽取：
      Round 1: 用基礎 prompt 抽 → validator → critic
      Round 2-3: 帶上一輪 corrections 重抽 → validator → critic
      任一輪 critic 通過 → 立即返回
      最多 MAX_CRITIC_ROUNDS 輪後返回最後一輪結果

    V4 多模態：multimodal=True 時，extraction 走 backend.call_multimodal（帶圖）；
    critic 仍走純文字 backend.call（critic 看 PDF 文字 + 答案就夠了，省成本）。

    V5 strict schema：啟用時 extraction 用 OPENAI_RESPONSE_FORMAT_STRICT；
    critic 仍用 json_object（critic 輸出簡單，不需要 strict）。

    回傳 (final_answer, history)
    """
    history: List[Dict] = []
    answer: Dict = {}
    critic_mode = get_critic_mode()
    num_images = len(images_b64) if images_b64 else 0
    extract_response_format = OPENAI_RESPONSE_FORMAT_STRICT if get_use_strict_schema() else None
    use_cot = get_use_cot()

    for round_num in range(1, MAX_CRITIC_ROUNDS + 1):
        # ---- 1. 組 prompt（第 1 輪用基礎，2+ 帶 feedback；multimodal / CoT 走不同 builder）----
        if multimodal and use_cot:
            if round_num == 1:
                prompt = build_extraction_prompt_multimodal_cot(pn, pdf_text, num_images)
            else:
                prev_corrections = history[-1]["corrections"]
                prompt = build_extraction_prompt_with_feedback_multimodal_cot(
                    pn, pdf_text, answer, prev_corrections, round_num, num_images
                )
        elif multimodal:
            if round_num == 1:
                prompt = build_extraction_prompt_multimodal(pn, pdf_text, num_images)
            else:
                prev_corrections = history[-1]["corrections"]
                prompt = build_extraction_prompt_with_feedback_multimodal(
                    pn, pdf_text, answer, prev_corrections, round_num, num_images
                )
        else:
            if round_num == 1:
                prompt = build_extraction_prompt(pn, pdf_text)
            else:
                prev_corrections = history[-1]["corrections"]
                prompt = build_extraction_prompt_with_feedback(
                    pn, pdf_text, answer, prev_corrections, round_num
                )

        # ---- 2. Extraction Agent ----
        t0 = time.time()
        try:
            if multimodal:
                raw = backend.call_multimodal(
                    prompt, images_b64, timeout=config.TIMEOUT,
                    response_format=extract_response_format,
                )
                answer = extract_json(raw)
            else:
                raw = backend.call(
                    prompt, timeout=config.TIMEOUT,
                    response_format=extract_response_format,
                )
                answer = extract_json(raw)
        except Exception as e:
            # JSON parse 失敗時走原本的 retry 機制（不帶 multimodal）
            print(f"      [round {round_num} extract retry] {type(e).__name__}: {e}", flush=True)
            answer = call_with_retry(backend, prompt)
            raw = json.dumps(answer, ensure_ascii=False)  # 給 logger 用 placeholder
        extract_elapsed_ms = int((time.time() - t0) * 1000)

        if logger:
            logger.log_extraction(
                round_num=round_num,
                prompt_chars=len(prompt),
                response_raw=raw,
                parsed_answer=answer,
                elapsed_ms=extract_elapsed_ms,
                multimodal=multimodal,
                num_images_sent=num_images,
            )

        # ---- 3. 本地 validator 修明顯錯（mV→V、L/W/H 範圍等）----
        answer, fixes = validate_and_fix(answer, pn)
        for fix in fixes:
            print(f"        [round {round_num} validator] {fix}", flush=True)
        if logger:
            logger.log_validator(round_num=round_num, fixes=fixes, answer_after=answer)

        # ---- 4. QA Critic Agent（V14: 可帶圖 if USE_MULTIMODAL_CRITIC=1）----
        is_passed, corrections, critic_debug = call_critic(
            backend, pn, pdf_text, answer,
            images_b64=images_b64 if multimodal else None,
        )
        history.append({
            "round": round_num,
            "fixes": fixes,
            "critic_passed": is_passed,
            "corrections": corrections,
        })
        if logger:
            logger.log_critic(
                round_num=round_num,
                mode=critic_debug["mode"],
                prompt_chars=critic_debug["prompt_chars"],
                response_raw=critic_debug["response_raw"],
                is_passed=is_passed,
                corrections=corrections,
                elapsed_ms=critic_debug["elapsed_ms"],
            )

        if is_passed:
            print(f"        [round {round_num}] critic PASS", flush=True)
            if logger:
                logger.log_final(total_rounds=round_num, final_answer=answer, critic_final_passed=True)
            return answer, history

        print(f"        [round {round_num}] critic FAIL, {len(corrections)} corrections", flush=True)
        for c in corrections[:3]:  # 只印前 3 條,避免洗版
            print(f"          · {c[:120]}", flush=True)

    # 跑完 MAX_CRITIC_ROUNDS 還沒通過,返回最後一輪結果
    print(f"        [exhausted] {MAX_CRITIC_ROUNDS} rounds done, using last answer", flush=True)
    if logger:
        logger.log_final(total_rounds=MAX_CRITIC_ROUNDS, final_answer=answer, critic_final_passed=False)
    return answer, history


# ===================================================================
# 3. 比對(數值容差 / 字串完全 / token 集合)
# ===================================================================

def load_specbook(path: Path) -> Dict[str, Dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    spec = {}
    for r in rows[1:]:
        if not r or r[0] is None:
            continue
        rec = dict(zip(headers, r))
        pn = str(rec.get("Part Number", "")).strip()
        if pn:
            spec[pn] = rec
    return spec


def _norm_token(t: str) -> str:
    s = t.lower()
    s = s.replace("μ", "u").replace("µ", "u")
    s = re.sub(r"\b(ta|tj|vr|ir|if|vf|tc|tamb)\s*=\s*\+?", "", s)
    s = re.sub(r"\b(ma|a|v|w)(dc|ac|pk|rms)\b", r"\1", s)
    s = s.replace(",", "").replace("，", "").replace("、", "").replace("+", "").replace("°", "")
    s = re.sub(r"\s+", "", s)

    def strip0(m):
        n = m.group(0)
        if "." in n:
            n = n.rstrip("0").rstrip(".")
        return n

    s = re.sub(r"\d+\.\d+", strip0, s)
    return s


def split_text(v) -> set:
    if v is None or not isinstance(v, str):
        return set()
    return {_norm_token(p) for p in re.split(r"[、]", v) if p.strip()}


def cmp_numeric(x, e) -> bool:
    if x is None or e is None:
        return x == e
    try:
        ef, xf = float(e), float(x)
    except (ValueError, TypeError):
        return str(x).strip() == str(e).strip()
    if ef == 0:
        return abs(xf) < 1e-6
    return abs(xf - ef) / abs(ef) <= config.NUMERIC_TOLERANCE


def cmp_exact(x, e) -> bool:
    if x is None or e is None:
        return x == e

    def norm(v):
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()

    return norm(x) == norm(e)


def cmp_text(x, e) -> Tuple[bool, float]:
    es = split_text(e)
    xs = split_text(x)
    if not es:
        return (not xs), (1.0 if not xs else 0.0)
    if not xs:
        return False, 0.0
    score = len(es & xs) / len(es)
    return score >= config.TEXT_MATCH_THRESHOLD, score


def evaluate(extracted: Dict, spec: Dict) -> Dict:
    out = {
        "overall": {"correct": 0, "total": 0, "accuracy": 0.0},
        "per_part": {},
        "cells": [],
    }
    for pn, exp_rec in spec.items():
        e = extracted.get(pn, {})
        pr = {"correct": 0, "total": 0, "accuracy": 0.0, "fields": {}}
        if "_error" in e:
            for f in config.FIELDS:
                d = {"expected": exp_rec.get(f), "extracted": None,
                     "match": False, "reason": "extraction failed"}
                pr["fields"][f] = d
                out["cells"].append({
                    "part": pn, "field": f,
                    "expected": exp_rec.get(f), "extracted": None, "match": False,
                })
            pr["total"] = len(config.FIELDS)
        else:
            for f in config.FIELDS:
                exp = exp_rec.get(f)
                ext = e.get(f)
                if f in config.NUMERIC_FIELDS:
                    ok = cmp_numeric(ext, exp)
                    d = {"expected": exp, "extracted": ext, "match": ok}
                elif f in config.TEXT_FIELDS:
                    ok, score = cmp_text(ext, exp)
                    d = {"expected": exp, "extracted": ext, "match": ok,
                         "token_score": round(score, 3)}
                else:
                    ok = cmp_exact(ext, exp)
                    d = {"expected": exp, "extracted": ext, "match": ok}
                pr["fields"][f] = d
                pr["total"] += 1
                if ok:
                    pr["correct"] += 1
                out["cells"].append({
                    "part": pn, "field": f,
                    "expected": exp, "extracted": ext, "match": ok,
                })
        pr["accuracy"] = pr["correct"] / pr["total"] if pr["total"] else 0
        out["per_part"][pn] = pr
        out["overall"]["correct"] += pr["correct"]
        out["overall"]["total"] += pr["total"]
    if out["overall"]["total"]:
        out["overall"]["accuracy"] = out["overall"]["correct"] / out["overall"]["total"]
    return out


# ===================================================================
# 4. Excel 輸出(3 個 sheet:Results / Comparison / Summary)
# ===================================================================

def write_excel(extracted: Dict, report: Dict, path: Path) -> None:
    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    green = PatternFill("solid", fgColor="D5F0D5")
    red = PatternFill("solid", fgColor="FBD5D5")

    # Sheet 1: Results
    ws1 = wb.active
    ws1.title = "Results"
    ws1.append(config.FIELDS)
    for c in ws1[1]:
        c.font = bold
        c.fill = header_fill
    for pn in config.PART_TO_PDF.keys():
        rec = extracted.get(pn, {})
        if "_error" in rec:
            row = [pn] + [None] * (len(config.FIELDS) - 1)
        else:
            row = [rec.get(f) for f in config.FIELDS]
            if not row[0]:
                row[0] = pn
            row = [v if isinstance(v, (str, int, float)) or v is None else str(v) for v in row]
        ws1.append(row)
    for i, f in enumerate(config.FIELDS, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = max(14, min(45, len(f) + 2))
    ws1.freeze_panes = "B2"

    # Sheet 2: Comparison
    ws2 = wb.create_sheet("Comparison")
    ws2.append(["Part Number", "Field", "Expected", "Extracted", "Match"])
    for c in ws2[1]:
        c.font = bold
        c.fill = header_fill
    for cell in report["cells"]:
        row = [
            cell["part"],
            cell["field"],
            "" if cell["expected"] is None else str(cell["expected"]),
            "" if cell["extracted"] is None else str(cell["extracted"]),
            "✓" if cell["match"] else "✗",
        ]
        ws2.append(row)
        cur = ws2.max_row
        fill = green if cell["match"] else red
        for col in range(1, 6):
            ws2.cell(row=cur, column=col).fill = fill
    for i, w in enumerate([14, 38, 50, 50, 8], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # Sheet 3: Summary
    ws3 = wb.create_sheet("Summary")
    ws3.append(["Part Number", "Correct", "Total", "Accuracy"])
    for c in ws3[1]:
        c.font = bold
        c.fill = header_fill
    for pn, pr in report["per_part"].items():
        ws3.append([pn, pr["correct"], pr["total"], f"{pr['accuracy'] * 100:.1f}%"])
    o = report["overall"]
    ws3.append([])
    ws3.append(["Overall", o["correct"], o["total"], f"{o['accuracy'] * 100:.1f}%"])
    last = ws3.max_row
    for col in range(1, 5):
        ws3.cell(row=last, column=col).font = bold
        ws3.cell(row=last, column=col).fill = PatternFill("solid", fgColor="FFE699")
    for i, w in enumerate([14, 10, 8, 12], start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"

    wb.save(path)


# ===================================================================
# 5. 自動歸檔
# ===================================================================

def get_cell_id(extractor: PdfExtractor, backend: LLMBackend) -> str:
    """生成統一 cell identifier 給 archive / iteration_log 用。
    V8 旗標 (env: VERSION_LABEL=v8) → v8_ 前綴 (拔 overfit + 通用化)
    V6 CoT 啟用 → v6_cot_ 前綴
    V5 strict schema 啟用 → v5_ 前綴
    multimodal → v4_
    其他 → v3_dual_agent_
    """
    critic_mode = get_critic_mode()
    label = os.environ.get("VERSION_LABEL", "").lower().strip()
    if label:
        version_prefix = label
    elif get_use_cot():
        version_prefix = "v6_cot"
    elif get_use_strict_schema():
        version_prefix = "v5_strict_schema"
    elif isinstance(extractor, MultiModalExtractor):
        version_prefix = "v4"
    else:
        version_prefix = "v3_dual_agent"
    return f"{version_prefix}_{extractor.name}_{backend.archive_tag}_{critic_mode}"


def archive_run(backend: LLMBackend, extractor: PdfExtractor, report: Dict) -> Path:
    """歸檔到 output/_archive/<cell_id>/"""
    multimodal = isinstance(extractor, MultiModalExtractor)
    archive_dir = config.ARCHIVE_DIR / get_cell_id(extractor, backend)
    archive_dir.mkdir(parents=True, exist_ok=True)
    # 拷檔
    if get_use_cot():
        excel_fname = "results_v6.xlsx"
    elif get_use_strict_schema():
        excel_fname = "results_v5.xlsx"
    elif multimodal:
        excel_fname = "results_v4.xlsx"
    else:
        excel_fname = "results_v3.xlsx"
    for fname in (excel_fname, "results.json"):
        src = config.OUTPUT_DIR / fname
        if src.exists():
            shutil.copy2(src, archive_dir / fname)
    # 也存一份 errors_detail.json 方便比對
    errors = [
        {"part": c["part"], "field": c["field"],
         "expected": str(c["expected"]) if c["expected"] is not None else "",
         "extracted": str(c["extracted"]) if c["extracted"] is not None else ""}
        for c in report["cells"] if not c["match"]
    ]
    (archive_dir / "errors_detail.json").write_text(
        json.dumps(errors, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    # 寫 summary.txt
    o = report["overall"]
    summary = [
        f"PDF Extractor: {extractor.name}",
        f"LLM Backend:   {backend.name}",
        f"Overall: {o['correct']}/{o['total']} = {o['accuracy']*100:.1f}%",
        "",
        "Per-part:",
    ]
    for pn, pr in report["per_part"].items():
        summary.append(f"  {pn:<15} {pr['correct']:>3}/{pr['total']:>3} = {pr['accuracy']*100:>5.1f}%")
    (archive_dir / "summary.txt").write_text("\n".join(summary), encoding="utf-8")
    return archive_dir


# ===================================================================
# 6. 主流程
# ===================================================================

def run() -> Dict:
    config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    backend = get_backend()
    extractor = get_pdf_extractor()
    critic_mode = get_critic_mode()
    cell_id = get_cell_id(extractor, backend)
    multimodal = isinstance(extractor, MultiModalExtractor)
    use_strict_schema = get_use_strict_schema()
    use_cot = get_use_cot()
    print(f"[backend]      {backend.name}")
    print(f"[pdf extractor] {extractor.name}")
    print(f"[critic mode]   {critic_mode}")
    print(f"[multimodal]    {multimodal}")
    print(f"[strict schema] {use_strict_schema}")
    print(f"[chain-of-thought] {use_cot}")
    print(f"[cell id]       {cell_id}")

    # V4 詳細迭代記錄器（V3 也會跑，不會壞既有流程）
    logger = IterationLogger(config.OUTPUT_DIR / "iteration_logs", cell_id)

    print("[1/4] 載入標準答案")
    spec = load_specbook(config.SPECBOOK_PATH)
    print(f"      共 {len(spec)} 顆元件")

    print("[2/4] PDF 抽取 + LLM 推理")
    extracted: Dict[str, Dict] = {}
    for pn, pdf_filename in config.PART_TO_PDF.items():
        pdf_path = config.DATASHEETS_DIR / pdf_filename
        if not pdf_path.exists():
            print(f"      [WARN] {pn}: 找不到 {pdf_filename}")
            extracted[pn] = {"_error": "PDF not found"}
            continue
        print(f"      - {pn}", flush=True)
        logger.start_part(pn)
        try:
            # ---- PDF 抽取（multimodal 多回傳 images）----
            t0 = time.time()
            if multimodal:
                pdf_text, images_b64, image_pages = extractor.extract_with_images(pdf_path)
            else:
                pdf_text = extractor.extract(pdf_path)
                images_b64 = None
                image_pages = []
            pdf_elapsed_ms = int((time.time() - t0) * 1000)
            logger.log_pdf_extract(
                extractor_name=extractor.name,
                text_chars=len(pdf_text),
                num_images=len(images_b64) if images_b64 else 0,
                elapsed_ms=pdf_elapsed_ms,
                image_pages=image_pages,
            )
            if multimodal:
                print(f"        [pdf] {len(pdf_text)} chars + {len(images_b64)} images (pages {image_pages})", flush=True)

            # ---- 雙 Agent 迭代 ----
            answer, history = run_dual_agent_loop(
                backend, pn, pdf_text,
                logger=logger,
                images_b64=images_b64,
                multimodal=multimodal,
            )
            extracted[pn] = answer
            extracted[pn]["_v3_history"] = history
        except Exception as e:
            print(f"        [ERROR] {e}")
            extracted[pn] = {"_error": str(e)}
            try:
                logger.log_error(f"{type(e).__name__}: {e}")
            except Exception:
                pass

    logger.close()

    (config.OUTPUT_DIR / "results.json").write_text(
        json.dumps(extracted, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print("[3/4] 評估比對")
    report = evaluate(extracted, spec)

    print("[4/4] 寫出 Excel")
    if use_cot:
        excel_filename = "results_v6.xlsx"
    elif use_strict_schema:
        excel_filename = "results_v5.xlsx"
    elif multimodal:
        excel_filename = "results_v4.xlsx"
    else:
        excel_filename = "results_v3.xlsx"
    excel_path = config.OUTPUT_DIR / excel_filename
    write_excel(extracted, report, excel_path)
    print(f"      → {excel_path}")

    print("[5/5] 自動歸檔")
    archive = archive_run(backend, extractor, report)
    print(f"      → {archive}")

    o = report["overall"]
    print()
    print(f"=== {extractor.name} × {backend.name} 整體準確率：{o['correct']}/{o['total']} = {o['accuracy'] * 100:.1f}% ===")
    print()
    print("各元件：")
    for pn, pr in report["per_part"].items():
        print(f"  {pn:<15} {pr['correct']:>3}/{pr['total']:>3} = {pr['accuracy'] * 100:>5.1f}%")
    return report


if __name__ == "__main__":
    try:
        run()
    except KeyboardInterrupt:
        print("\nUser interrupt")
        sys.exit(130)
